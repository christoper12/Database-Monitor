import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from sqlalchemy import create_engine
from openpyxl.utils import get_column_letter
import time
import os
import configparser

# =========================
# EXPORT FUNCTION
# =========================
def export_excel_file(open_file=False):

    # =========================
    # LOAD CONFIG
    # =========================
    config = configparser.ConfigParser()
    config.read('config.ini')

    HOST = config.get('MYSQL', 'HOST')
    USER = config.get('MYSQL', 'USER')
    PASSWORD = config.get('MYSQL', 'PASSWORD')
    PORT = config.getint('MYSQL', 'PORT')

    print("=" * 60)
    print("START MYSQL DATABASE EXPORT")
    print("=" * 60)

    start_time = time.time()

    # =========================
    # CONNECT MYSQL
    # =========================
    print("[1/6] Connecting to MySQL...")

    engine = create_engine(
        f"mysql+pymysql://{USER}:{PASSWORD}@{HOST}:{PORT}"
    )

    print("SUCCESS: Connected to MySQL")

    # =========================
    # GET DATABASE LIST
    # =========================
    print("[2/6] Fetching database list...")

    db_query = """
    SELECT SCHEMA_NAME
    FROM information_schema.SCHEMATA
    WHERE SCHEMA_NAME NOT IN (
        'information_schema',
        'performance_schema',
        'mysql',
        'sys'
    )
    ORDER BY SCHEMA_NAME;
    """

    databases = pd.read_sql(db_query, engine)

    total_db = len(databases)

    print(f"SUCCESS: Found {total_db} databases")

    # =========================
    # CREATE EXCEL
    # =========================
    print("[3/6] Creating Excel workbook...")

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    print("SUCCESS: Workbook created")

    # =========================
    # MASTER DATA
    # =========================
    master_data = []

    # =========================
    # LOOP DATABASES
    # =========================
    print("[4/6] Processing databases...")
    print("-" * 60)

    for index, db in enumerate(databases['SCHEMA_NAME'], start=1):

        db_start = time.time()

        print(f"[{index}/{total_db}] Processing DB: {db}")

        query = f"""
        SELECT
          TABLE_NAME AS `Table`,
          ROUND((DATA_LENGTH + INDEX_LENGTH) / 1024 / 1024, 2) AS `Size (MB)`
        FROM
          information_schema.TABLES
        WHERE
          TABLE_SCHEMA = '{db}'
        ORDER BY
          (DATA_LENGTH + INDEX_LENGTH) DESC;
        """

        total_query = f"""
        SELECT
          ROUND(
            SUM(DATA_LENGTH + INDEX_LENGTH)
            / 1024 / 1024 / 1024,
            4
          ) AS total_size_gb
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = '{db}';
        """

        try:

            # =========================
            # TABLE DETAIL
            # =========================
            df = pd.read_sql(query, engine)

            total_tables = len(df)

            print(f"     Tables found : {total_tables}")

            for i, table_name in enumerate(df['Table'], start=1):
                print(f"        [{i}/{total_tables}] {table_name}")

            # =========================
            # DATABASE TOTAL SIZE
            # =========================
            total_df = pd.read_sql(total_query, engine)

            total_size_gb = total_df.iloc[0]['total_size_gb']

            if pd.isna(total_size_gb):
                total_size_gb = 0

            master_data.append({
                "Database": db,
                "Total Size (GB)": round(float(total_size_gb), 4),
                "Total Tables": total_tables
            })

            # =========================
            # CREATE SHEET
            # =========================
            ws = wb.create_sheet(title=db[:31])

            # Header
            ws.append([f"Database: {db}"])
            ws.append([f"Total Size (GB): {round(float(total_size_gb), 4)}"])

            # Back to master link
            back_cell = ws.cell(
                row=3,
                column=1,
                value="← Back to MASTER"
            )

            back_cell.hyperlink = "#'MASTER'!A1"
            back_cell.style = "Hyperlink"

            ws.append([])

            # Insert Data
            for row in dataframe_to_rows(
                df,
                index=False,
                header=True
            ):
                ws.append(row)

            elapsed = round(
                time.time() - db_start,
                2
            )

            print(f"     SUCCESS ({elapsed} sec)")
            print("-" * 60)

        except Exception as e:

            print(f"     ERROR: {e}")
            print("-" * 60)

    # =========================
    # CREATE MASTER SHEET
    # =========================
    print("[5/6] Creating MASTER sheet...")

    master_df = pd.DataFrame(master_data)

    # Sort by biggest database
    master_df = master_df.sort_values(
        by="Total Size (GB)",
        ascending=False
    )

    ws_master = wb.create_sheet(
        title="MASTER",
        index=0
    )

    ws_master.append(["MYSQL DATABASE SUMMARY"])
    ws_master.append([])

    # Header
    headers = [
        "Database",
        "Total Size (GB)",
        "Total Tables"
    ]

    ws_master.append(headers)

    # Rows
    for _, row in master_df.iterrows():

        db_name = row["Database"]

        current_row = ws_master.max_row + 1

        # Database name
        cell = ws_master.cell(
            row=current_row,
            column=1,
            value=db_name
        )

        # Hyperlink to sheet
        cell.hyperlink = f"#'{db_name[:31]}'!A1"

        # Style hyperlink
        cell.style = "Hyperlink"

        # Other columns
        ws_master.cell(
            row=current_row,
            column=2,
            value=row["Total Size (GB)"]
        )

        ws_master.cell(
            row=current_row,
            column=3,
            value=row["Total Tables"]
        )

    print("SUCCESS: MASTER sheet created")

    # =========================
    # SAVE FILE
    # =========================
    print("[6/6] Saving Excel file...")

    # Create output folder
    output_dir = "output"

    os.makedirs(output_dir, exist_ok=True)

    # Filename
    now = datetime.now().strftime(
        "%d-%m-%Y_%H-%M-%S"
    )

    output_file = os.path.join(
        output_dir,
        f"mysql_table_sizes_{now}.xlsx"
    )

    # =========================
    # AUTO SIZE COLUMNS
    # =========================
    print("Auto sizing columns...")

    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = 0

            column = column_cells[0].column
            column_letter = get_column_letter(column)

            for cell in column_cells:
                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except:
                    pass

            adjusted_width = max_length + 4
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save
    wb.save(output_file)

    # Open file if requested
    if open_file:
        os.startfile(output_file)
        print("Opening Excel file...")

    total_time = round(
        time.time() - start_time,
        2
    )

    print("=" * 60)
    print("SUCCESS: File saved")
    print(f"FILE    : {output_file}")
    print(f"DURATION: {total_time} seconds")
    print("=" * 60)

    # Return file path
    return output_file


# =========================
# RUN MANUALLY
# =========================
if __name__ == "__main__":

    open_excel = input(
        "Open Excel file now? (y/n): "
    )

    result = export_excel_file(
        open_file=(open_excel.lower() == "y")
    )

    print(f"Generated: {result}")